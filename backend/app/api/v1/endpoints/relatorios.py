from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import case, func, inspect, literal, or_, select, text
from sqlalchemy.orm import Session

from app.core.org_context import get_current_org
from app.core.security import require_roles
from app.db.session import get_db
from app.models.company import Company
from app.models.company_profile import CompanyProfile
from app.models.company_tax import CompanyTax
from app.models.org import Org

router = APIRouter()

MANDATORY_FIELDS = ("id", "cnpj", "razao_social")
TABLE_COMPANIES = "companies"
TABLE_PROFILES = "company_profiles"
SENSITIVE_RE = re.compile(r"(senha|password|token|hash)", re.IGNORECASE)
INTERNAL_COMPANY_FIELDS = {"org_id"}
INTERNAL_PROFILE_FIELDS = {"id", "org_id", "company_id", "raw"}

FIELD_LABEL_OVERRIDES = {
    "id": "ID",
    "cnpj": "CNPJ",
    "razao_social": "Razao Social",
    "nome_fantasia": "Nome Fantasia",
    "is_active": "Status",
    "cpf": "CPF Responsavel Legal",
    "company_cpf": "Cadastros PF",
    "fs_dirname": "Apelido (pasta)",
    "possui_debitos": "Possui debitos",
    "sem_debitos": "Sem debitos",
}


class RelatorioExportRequest(BaseModel):
    campos: list[str] = Field(default_factory=list)


@dataclass(frozen=True)
class ExportFieldDef:
    key: str
    selectable: object
    label: str


def _is_sensitive_or_empty(field_name: str) -> bool:
    normalized = str(field_name or "").strip()
    if not normalized:
        return True
    return bool(SENSITIVE_RE.search(normalized))


def _list_columns(db: Session, table_name: str) -> list[str]:
    dialect = db.bind.dialect.name if db.bind else ""
    if dialect == "postgresql":
        rows = db.execute(
            text(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = :schema AND table_name = :table_name
                ORDER BY ordinal_position
                """
            ),
            {"schema": "public", "table_name": table_name},
        ).scalars()
        return [str(name) for name in rows]

    inspector = inspect(db.bind)
    return [str(col["name"]) for col in inspector.get_columns(table_name)]


def _humanize_field_name(field_name: str) -> str:
    if field_name in FIELD_LABEL_OVERRIDES:
        return FIELD_LABEL_OVERRIDES[field_name]
    return str(field_name or "").strip().replace("_", " ").title()


def _build_has_debito_expr():
    monitored = [
        CompanyTax.taxa_funcionamento,
        CompanyTax.taxa_publicidade,
        CompanyTax.taxa_vig_sanitaria,
        CompanyTax.taxa_localiz_instalacao,
        CompanyTax.taxa_ocup_area_publica,
        CompanyTax.tpi,
        CompanyTax.status_taxas,
    ]
    checks = [func.lower(func.coalesce(column, "")).like("%aberto%") for column in monitored]
    return or_(*checks)


def _build_allowed_field_map(db: Session) -> dict[str, ExportFieldDef]:
    allowed: dict[str, ExportFieldDef] = {}

    for column in _list_columns(db, TABLE_COMPANIES):
        if _is_sensitive_or_empty(column) or column in INTERNAL_COMPANY_FIELDS:
            continue
        output_name = "company_cpf" if column == "cpf" else column
        allowed[output_name] = ExportFieldDef(
            key=output_name,
            selectable=getattr(Company, column),
            label=_humanize_field_name(output_name),
        )

    for column in _list_columns(db, TABLE_PROFILES):
        if _is_sensitive_or_empty(column) or column in INTERNAL_PROFILE_FIELDS:
            continue
        if column in allowed:
            continue
        allowed[column] = ExportFieldDef(
            key=column,
            selectable=getattr(CompanyProfile, column),
            label=_humanize_field_name(column),
        )

    has_debito = _build_has_debito_expr()
    allowed["possui_debitos"] = ExportFieldDef(
        key="possui_debitos",
        selectable=case((has_debito, literal("Sim")), else_=literal("Nao")),
        label=_humanize_field_name("possui_debitos"),
    )
    allowed["sem_debitos"] = ExportFieldDef(
        key="sem_debitos",
        selectable=case((has_debito, literal("Nao")), else_=literal("Sim")),
        label=_humanize_field_name("sem_debitos"),
    )

    for field in MANDATORY_FIELDS:
        allowed[field] = ExportFieldDef(
            key=field,
            selectable=getattr(Company, field),
            label=_humanize_field_name(field),
        )
    return allowed


def _normalize_requested_fields(campos: list[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in campos or []:
        candidate = str(raw or "").strip()
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        normalized.append(candidate)
    return normalized


def _format_field_value(field: str, value: object) -> object:
    if value is None:
        return ""
    if field == "is_active":
        return "Ativa" if bool(value) else "Inativa"
    if isinstance(value, bool):
        return "Sim" if value else "Nao"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _build_export_fields(requested: list[str], allowed_fields: dict[str, ExportFieldDef]) -> list[str]:
    invalid = [field for field in requested if field not in allowed_fields]
    if invalid:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "Campos invalidos para exportacao",
                "invalidos": invalid,
            },
        )
    ordered = list(MANDATORY_FIELDS)
    for field in requested:
        if field not in ordered:
            ordered.append(field)
    return ordered


def _format_workbook(export_fields: list[str], allowed_fields: dict[str, ExportFieldDef], rows: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"

    headers = [allowed_fields[field].label for field in export_fields]
    ws.append(headers)
    for row in rows:
        ws.append([_format_field_value(field, row.get(field)) for field in export_fields])

    header_fill_default = PatternFill(fill_type="solid", start_color="1F3864", end_color="1F3864")
    header_fill_red = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")
    header_fill_green = PatternFill(fill_type="solid", start_color="2E7D32", end_color="2E7D32")
    cell_fill_red = PatternFill(fill_type="solid", start_color="FDECEC", end_color="FDECEC")
    cell_fill_green = PatternFill(fill_type="solid", start_color="E8F5E9", end_color="E8F5E9")
    odd_row_fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    even_row_fill = PatternFill(fill_type="solid", start_color="EEF2F7", end_color="EEF2F7")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_index, cell in enumerate(ws[1], start=1):
        field_key = export_fields[col_index - 1]
        if field_key == "possui_debitos":
            cell.fill = header_fill_red
        elif field_key == "sem_debitos":
            cell.fill = header_fill_green
        else:
            cell.fill = header_fill_default
        cell.font = header_font
        cell.border = border

    for row_index in range(2, ws.max_row + 1):
        base_fill = even_row_fill if row_index % 2 == 0 else odd_row_fill
        for col_index in range(1, ws.max_column + 1):
            field_key = export_fields[col_index - 1]
            cell = ws.cell(row=row_index, column=col_index)
            cell.border = border

            if field_key in {"possui_debitos", "sem_debitos"}:
                if str(cell.value).strip().lower() == "sim":
                    cell.fill = cell_fill_red if field_key == "possui_debitos" else cell_fill_green
                else:
                    cell.fill = base_fill
            else:
                cell.fill = base_fill

    for col_index in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_index)
        max_len = 0
        for row_index in range(1, ws.max_row + 1):
            value = ws.cell(row=row_index, column=col_index).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[col_letter].width = max(15, min(60, max_len + 2))

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    return wb


@router.get("/campos")
def list_report_fields(
    db: Session = Depends(get_db),
    _org: Org = Depends(get_current_org),
    _user=Depends(require_roles("ADMIN", "DEV", "VIEW")),
) -> dict:
    allowed_fields = _build_allowed_field_map(db)
    optionals = [field for field in allowed_fields if field not in MANDATORY_FIELDS]
    labels = {field: allowed_fields[field].label for field in allowed_fields}
    return {"obrigatorios": list(MANDATORY_FIELDS), "opcionais": optionals, "labels": labels}


@router.post("/exportar")
def export_report(
    payload: RelatorioExportRequest,
    db: Session = Depends(get_db),
    org: Org = Depends(get_current_org),
    _user=Depends(require_roles("ADMIN", "DEV", "VIEW")),
):
    allowed_fields = _build_allowed_field_map(db)
    requested = _normalize_requested_fields(payload.campos)
    export_fields = _build_export_fields(requested, allowed_fields)

    selected_columns = [allowed_fields[field].selectable.label(field) for field in export_fields]
    include_pf_registers = "company_cpf" in export_fields

    statement = (
        select(*selected_columns)
        .select_from(Company)
        .outerjoin(
            CompanyProfile,
            (Company.id == CompanyProfile.company_id) & (Company.org_id == CompanyProfile.org_id),
        )
        .outerjoin(
            CompanyTax,
            (Company.id == CompanyTax.company_id) & (Company.org_id == CompanyTax.org_id),
        )
        .where(Company.org_id == org.id)
        .order_by(Company.created_at.desc())
    )
    if not include_pf_registers:
        statement = statement.where(Company.cnpj.is_not(None))

    rows = db.execute(statement).mappings().all()
    workbook = _format_workbook(export_fields, allowed_fields, [dict(row) for row in rows])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"relatorio_eControle_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
