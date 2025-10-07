"""
repo_excel.py - Repositório Excel com leitura/escrita segura
Compatível com config.yaml (uma aba PROCESSOS com múltiplas Tabelas/ListObjects)

Principais recursos:
- Lock de arquivo (portalocker) para leitura/escrita segura de .xlsm
- Mapeamento por aliases definidos no config.yaml
- Leitura por ABA (cabeçalho heurístico) **e** por TABELA (ListObject) usando openpyxl.tables
- Suporte a table_names → ex.: aba PROCESSOS com as tabelas Diversos, Funcionamento, etc.
- Utilitários para diagnosticar colunas (build_column_map*)
"""
from __future__ import annotations

import logging
import re
import time
import unicodedata
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set

import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries
import portalocker
import yaml

logger = logging.getLogger(__name__)


class ExcelRepo:
    """Repositório para leitura/escrita segura em Excel com macros"""

    def __init__(self, excel_path: str, config_path: str = "config.yaml", *, data_only: bool = True):
        self.excel_path = Path(excel_path)
        self.config = self._load_config(config_path)
        self.wb: Optional[openpyxl.Workbook] = None
        # mapas de colunas (por chave lógica)
        self._column_maps: Dict[str, Dict[str, int]] = {}
        self._lock: Optional[portalocker.Lock] = None
        self._data_only = data_only

    # ------------------------------------------------------------------
    # Config
    # ------------------------------------------------------------------
    def _load_config(self, config_path: str) -> dict:
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                cfg = yaml.safe_load(f) or {}
        except FileNotFoundError:
            logger.warning("Config não encontrado: %s, usando padrões", config_path)
            cfg = {}
        # sane defaults
        cfg.setdefault("sheet_names", {})
        cfg.setdefault("table_names", {})
        cfg.setdefault("column_aliases", {})
        return cfg

    # ------------------------------------------------------------------
    # Header detection & mapping (ABA ou TABELA)
    # ------------------------------------------------------------------
    def _find_header_row(self, ws: Worksheet, max_rows: int = 10) -> int:
        """Heurística para ABAs sem tabela: primeira linha com >=4 valores e palavras-chave."""
        keywords = {"ID", "EMPRESA", "CNPJ", "PROTOCOLO", "TIPO", "STATUS", "MUNIC"}
        for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
            row_values = [cell.value for cell in ws[row_idx]]
            non_null = [v for v in row_values if v is not None and str(v).strip()]
            if len(non_null) >= 4:
                row_upper = [str(v).upper().strip() for v in non_null]
                if any(kw in " ".join(row_upper) for kw in keywords):
                    logger.debug("Cabeçalho encontrado na linha %s", row_idx)
                    return row_idx
        logger.warning("Cabeçalho não encontrado, usando linha 2 como padrão")
        return 2

    def _normalize_alias(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip().strip("'\"")
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        return text.upper()

    def _canonical_from_header(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip().strip("'\"")
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = re.sub(r"[^0-9A-Z]+", "_", text.upper())
        return text.strip("_")

    def _build_column_map_from_header_cells(self, header_cells: Iterable[Any], sheet_key: str) -> Dict[str, int]:
        aliases = self.config.get("column_aliases", {}).get(sheet_key, {})
        normalized_aliases: Dict[str, Set[str]] = {
            canonical: {self._normalize_alias(alias) for alias in alias_list}
            for canonical, alias_list in aliases.items()
        }
        col_map: Dict[str, int] = {}
        for idx, cell in enumerate(header_cells, start=1):
            value = getattr(cell, "value", cell)
            if value is None:
                continue
            normalized_value = self._normalize_alias(value)
            matched = False
            for canonical, alias_set in normalized_aliases.items():
                if normalized_value and normalized_value in alias_set:
                    col_map[canonical] = idx
                    matched = True
                    break
            if not matched:
                fallback_key = self._canonical_from_header(value)
                if fallback_key and fallback_key not in col_map:
                    col_map[fallback_key] = idx
        if not col_map:
            logger.warning("Nenhuma coluna mapeada para %s", sheet_key)
        return col_map

    def _build_column_map_sheet(self, ws: Worksheet, sheet_key: str) -> Dict[str, int]:
        header_row_idx = self._find_header_row(ws)
        header_row = ws[header_row_idx]
        return self._build_column_map_from_header_cells(header_row, sheet_key)

    def _get_table_by_name(self, ws: Worksheet, table_name: str):
        """Retorna o objeto Table pelo nome (compatível com openpyxl>=3)."""
        tables = getattr(ws, "tables", None)
        if tables is None:
            # fallback versões antigas
            tables = {t.name: t for t in getattr(ws, "_tables", [])}
        if isinstance(tables, dict):
            return tables.get(table_name)
        # lista de tables (fallback raro)
        for t in list(tables or []):
            if getattr(t, "name", None) == table_name:
                return t
        return None

    def _iter_table_rows(self, ws: Worksheet, table) -> Iterable[List[Any]]:
        """Itera as linhas (como células) dentro do range da Tabela."""
        if table is None:
            return []
        ref = table.ref  # e.g., 'A1:K100'
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            yield r

    # Públicos (para diagnóstico)
    def build_column_map(self, sheet_name: str, sheet_key: str) -> Dict[str, int]:
        if not self.wb:
            raise RuntimeError("Workbook não está aberto")
        ws = self.wb[sheet_name]
        col_map = self._build_column_map_sheet(ws, sheet_key)
        self._column_maps[sheet_key] = col_map
        return col_map

    def build_column_map_table(self, sheet_name: str, table_name: str, sheet_key: str) -> Dict[str, int]:
        if not self.wb:
            raise RuntimeError("Workbook não está aberto")
        ws = self.wb[sheet_name]
        tbl = self._get_table_by_name(ws, table_name)
        if not tbl:
            raise ValueError(f"Tabela '{table_name}' não encontrada na aba '{sheet_name}'")
        rows = list(self._iter_table_rows(ws, tbl))
        if not rows:
            return {}
        header = rows[0]
        col_map = self._build_column_map_from_header_cells(header, sheet_key)
        self._column_maps[sheet_key] = col_map
        return col_map

    def get_column_map(self, sheet_key: str) -> Dict[str, int]:
        return self._column_maps.get(sheet_key, {})

    # ------------------------------------------------------------------
    # Open / Save / Close com lock
    # ------------------------------------------------------------------
    def open(self, lock_timeout: int = 5):
        import os
        retry_count = 0
        max_retries = 3
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Arquivo não encontrado: {self.excel_path}")
        while retry_count < max_retries:
            try:
                self._lock = portalocker.Lock(
                    str(self.excel_path), "r", timeout=lock_timeout, flags=portalocker.LOCK_SH
                )
                self._lock.acquire()
                self.wb = load_workbook(
                    self.excel_path,
                    keep_vba=True,
                    data_only=self._data_only,
                )
                logger.info("Excel aberto: %s", self.excel_path)
                return
            except (PermissionError, portalocker.exceptions.LockException) as e:
                retry_count += 1
                logger.warning("Erro ao abrir arquivo (tentativa %s/%s): %s", retry_count, max_retries, str(e))
                if retry_count >= max_retries:
                    raise RuntimeError(
                        "Não foi possível abrir o arquivo após tentativas. Verifique se não está aberto no Excel."
                    )
                time.sleep(1)

    def save(self):
        if not self.wb:
            raise RuntimeError("Workbook não está aberto")
        try:
            self.wb.save(self.excel_path)
            logger.info("Excel salvo com sucesso")
        except PermissionError:
            raise RuntimeError("Não foi possível salvar: arquivo aberto em outro programa.")

    def close(self):
        if self.wb:
            self.wb.close()
            self.wb = None
            self._column_maps.clear()
        if self._lock:
            try:
                self._lock.release()
            finally:
                self._lock = None

    # ------------------------------------------------------------------
    # Read / Write por ABA
    # ------------------------------------------------------------------
    def read_sheet(self, sheet_name: str, sheet_key: str = "") -> List[Dict[str, Any]]:
        if not self.wb:
            raise RuntimeError("Workbook não está aberto")
        if sheet_name not in self.wb.sheetnames:
            logger.error("Aba '%s' não encontrada", sheet_name)
            return []
        ws = self.wb[sheet_name]
        if sheet_key not in self._column_maps:
            self._column_maps[sheet_key] = self._build_column_map_sheet(ws, sheet_key)
        col_map = self._column_maps[sheet_key]
        if not col_map:
            logger.warning("Mapa de colunas vazio para %s", sheet_key)
            return []
        header_row_idx = self._find_header_row(ws)
        data: List[Dict[str, Any]] = []
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_data: Dict[str, Any] = {}
            is_empty = True
            for canonical, col_idx in col_map.items():
                value = ws.cell(row_idx, col_idx).value
                if value not in (None, ""):
                    is_empty = False
                row_data[canonical] = value if value is not None else ""
            if not is_empty:
                data.append(row_data)
        logger.info("Lidos %s registros da aba %s", len(data), sheet_name)
        return data

    # ------------------------------------------------------------------
    # Read por TABELA (ListObject) — chave para PROCESSOS & ÚTEIS
    # ------------------------------------------------------------------
    def read_table(self, sheet_name: str, table_name: str, sheet_key: str) -> List[Dict[str, Any]]:
        if not self.wb:
            raise RuntimeError("Workbook não está aberto")
        if sheet_name not in self.wb.sheetnames:
            logger.error("Aba '%s' não encontrada", sheet_name)
            return []
        ws = self.wb[sheet_name]
        tbl = self._get_table_by_name(ws, table_name)
        if not tbl:
            logger.error("Tabela '%s' não encontrada na aba '%s'", table_name, sheet_name)
            return []
        rows = list(self._iter_table_rows(ws, tbl))
        if not rows:
            return []
        header = rows[0]
        if sheet_key not in self._column_maps:
            self._column_maps[sheet_key] = self._build_column_map_from_header_cells(header, sheet_key)
        col_map = self._column_maps[sheet_key]
        if not col_map:
            logger.warning("Mapa de colunas vazio para %s (tabela %s)", sheet_key, table_name)
            return []
        data: List[Dict[str, Any]] = []
        for r in rows[1:]:  # pula header
            row_data: Dict[str, Any] = {}
            is_empty = True
            for canonical, idx in col_map.items():
                # idx é 1-based relativo à tabela
                cell = r[idx - 1]
                value = cell.value
                if value not in (None, ""):
                    is_empty = False
                row_data[canonical] = value if value is not None else ""
            if not is_empty:
                data.append(row_data)
        logger.info("Lidos %s registros da tabela %s/%s", len(data), sheet_name, table_name)
        return data

    # ------------------------------------------------------------------
    # Utilities específicos para a estrutura do config.yaml
    # ------------------------------------------------------------------
    def list_tables(self, sheet_name: str) -> List[str]:
        if not self.wb:
            raise RuntimeError("Workbook não está aberto")
        if sheet_name not in self.wb.sheetnames:
            return []
        ws = self.wb[sheet_name]
        tables = getattr(ws, "tables", None)
        if isinstance(tables, dict):
            return list(tables.keys())
        return [t.name for t in (tables or [])]


# Utilitário CLI para diagnóstico rápido
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    # Ajuste o caminho do arquivo abaixo antes de rodar localmente
    repo = ExcelRepo("G:/PMA/LISTA EMPRESAS - NETO CONTABILIDADE 2025.xlsm")
    repo.open()
    try:
        cfg = repo.config
        print("Sheets:", cfg.get("sheet_names"))
        print("Tables (PROCESSOS):", cfg.get("table_names", {}).get("processos", {}))
        # Abas simples
        for key, name in {
            "empresas": cfg["sheet_names"].get("empresas", "EMPRESAS"),
            "licencas": cfg["sheet_names"].get("licencas", "LICENÇAS"),
            "taxas": cfg["sheet_names"].get("taxas", "TAXAS"),
        }.items():
            print(f"=== {name} (sheet:{key}) ===")
            repo.build_column_map(name, key)
            print(repo.get_column_map(key))
        # Tabelas de PROCESSOS
        proc_sheet = cfg["sheet_names"].get("processos", "PROCESSOS")
        proc_tables = cfg.get("table_names", {}).get("processos", {})
        for key, tbl_name in proc_tables.items():
            sheet_key = f"processos_{key}"
            print(f"=== Tabela {tbl_name} (sheet_key:{sheet_key}) ===")
            try:
                repo.build_column_map_table(proc_sheet, tbl_name, sheet_key)
                print(repo.get_column_map(sheet_key))
            except Exception as e:
                print("[warn]", e)
    finally:
        repo.close()
