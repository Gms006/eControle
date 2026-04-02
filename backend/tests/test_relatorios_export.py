from io import BytesIO

from openpyxl import load_workbook


def _login(client, email: str, password: str) -> str:
    response = client.post(
        "/api/v1/auth/login",
        json={"email": email, "password": password},
    )
    assert response.status_code == 200
    return response.json()["access_token"]


def _create_company_cnpj(client, token: str, cnpj: str, razao_social: str) -> str:
    headers = {"Authorization": f"Bearer {token}"}
    create_response = client.post(
        "/api/v1/companies",
        headers=headers,
        json={
            "cnpj": cnpj,
            "razao_social": razao_social,
            "nome_fantasia": "Relatorio Ltda",
            "municipio": "Anapolis",
            "uf": "GO",
            "is_active": True,
        },
    )
    assert create_response.status_code == 200
    company_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v1/companies/{company_id}",
        headers=headers,
        json={
            "categoria": "Comercio",
            "cpf": "11122233344",
            "telefone": "(62) 99999-9999",
        },
    )
    assert patch_response.status_code == 200
    return company_id


def _create_company_cpf(client, token: str, cpf: str, razao_social: str) -> str:
    headers = {"Authorization": f"Bearer {token}"}
    create_response = client.post(
        "/api/v1/companies",
        headers=headers,
        json={
            "company_cpf": cpf,
            "razao_social": razao_social,
            "municipio": "Anapolis",
            "uf": "GO",
            "is_active": True,
        },
    )
    assert create_response.status_code == 200
    return create_response.json()["id"]


def test_campos_endpoint_returns_labels_and_optional_fields(client):
    token = _login(client, "admin@example.com", "admin123")
    headers = {"Authorization": f"Bearer {token}"}
    _create_company_cnpj(client, token, "11.222.333/0001-44", "Empresa Relatorio")

    response = client.get("/api/v1/relatorios/campos", headers=headers)
    assert response.status_code == 200
    payload = response.json()
    assert payload["obrigatorios"] == ["id", "cnpj", "razao_social"]
    assert "nome_fantasia" in payload["opcionais"]
    assert "company_cpf" in payload["opcionais"]
    assert "possui_debitos" in payload["opcionais"]
    assert payload["labels"]["nome_fantasia"] == "Nome Fantasia"
    assert payload["labels"]["is_active"] == "Status"
    assert payload["labels"]["cpf"] == "CPF Responsavel Legal"
    assert payload["labels"]["fs_dirname"] == "Apelido (pasta)"
    assert payload["labels"]["company_cpf"] == "Cadastros PF"


def test_export_endpoint_requires_valid_fields(client):
    token = _login(client, "admin@example.com", "admin123")
    headers = {"Authorization": f"Bearer {token}"}
    _create_company_cnpj(client, token, "22.333.444/0001-55", "Empresa Relatorio")

    response = client.post(
        "/api/v1/relatorios/exportar",
        headers=headers,
        json={"campos": ["nome_fantasia", "campo_inexistente"]},
    )
    assert response.status_code == 400
    payload = response.json()
    assert payload["detail"]["message"] == "Campos invalidos para exportacao"
    assert "campo_inexistente" in payload["detail"]["invalidos"]


def test_export_headers_are_formatted_and_include_debito_columns(client):
    token = _login(client, "admin@example.com", "admin123")
    headers = {"Authorization": f"Bearer {token}"}
    _create_company_cnpj(client, token, "33.444.555/0001-66", "Empresa Relatorio")

    response = client.post(
        "/api/v1/relatorios/exportar",
        headers=headers,
        json={"campos": ["nome_fantasia", "is_active", "cpf", "possui_debitos", "sem_debitos"]},
    )
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    headers_row = [cell.value for cell in ws[1]]
    assert headers_row[:3] == ["ID", "CNPJ", "Razao Social"]
    assert "Nome Fantasia" in headers_row
    assert "Status" in headers_row
    assert "CPF Responsavel Legal" in headers_row
    assert "Possui debitos" in headers_row
    assert "Sem debitos" in headers_row


def test_export_without_company_cpf_hides_pf_registers(client):
    token = _login(client, "admin@example.com", "admin123")
    headers = {"Authorization": f"Bearer {token}"}
    _create_company_cnpj(client, token, "44.555.666/0001-77", "Empresa PJ")
    _create_company_cpf(client, token, "123.456.789-01", "Pessoa Fisica")

    response_without_pf = client.post(
        "/api/v1/relatorios/exportar",
        headers=headers,
        json={"campos": ["nome_fantasia"]},
    )
    assert response_without_pf.status_code == 200
    ws_without_pf = load_workbook(filename=BytesIO(response_without_pf.content)).active
    data_rows_without_pf = list(ws_without_pf.iter_rows(min_row=2, values_only=True))
    assert len(data_rows_without_pf) == 1
    assert str(data_rows_without_pf[0][2]).lower() == "empresa pj".lower()

    response_with_pf = client.post(
        "/api/v1/relatorios/exportar",
        headers=headers,
        json={"campos": ["company_cpf"]},
    )
    assert response_with_pf.status_code == 200
    ws_with_pf = load_workbook(filename=BytesIO(response_with_pf.content)).active
    data_rows_with_pf = list(ws_with_pf.iter_rows(min_row=2, values_only=True))
    assert len(data_rows_with_pf) == 2
    headers_row = [cell.value for cell in ws_with_pf[1]]
    assert "Cadastros PF" in headers_row
