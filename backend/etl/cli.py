"""Command line interface for the ETL pipeline."""
from __future__ import annotations

import json
import os
import uuid
from pathlib import Path
from typing import Optional

import sqlalchemy as sa
import typer
from openpyxl import load_workbook
from typer.main import TyperArgument

from .contracts import load_contract
from .extract_xlsm import load as load_source
from .load_upsert import run as run_loader
from .transform_normalize import transform

from dotenv import load_dotenv
load_dotenv("backend/.env")


# Typer <0.9.0 defines TyperArgument.make_metavar(self) while
# Click>=8.1 invokes make_metavar(self, param). When those versions are
# combined Typer raises ``TypeError: TyperArgument.make_metavar() takes 1
# positional argument but 2 were given`` when the CLI is created.
# Patch TyperArgument at import time to tolerate the extra argument.
if TyperArgument.make_metavar.__code__.co_argcount == 1:  # pragma: no cover - defensive
    _original_make_metavar = TyperArgument.make_metavar

    def _compatible_make_metavar(self, ctx=None):
        """Bridge Typer's old signature with Click>=8.1 expectations."""

        if ctx is None:
            return _original_make_metavar(self)

        if self.metavar is not None:
            return self.metavar

        var = (self.name or "").upper()
        if not self.required:
            var = "[{}]".format(var)

        try:
            type_var = self.type.get_metavar(self, ctx)  # type: ignore[arg-type]
        except TypeError:  # fallback for exotic ParamTypes
            type_var = self.type.get_metavar(self)

        if type_var:
            var += f":{type_var}"
        if self.nargs != 1:
            var += "..."
        return var

    TyperArgument.make_metavar = _compatible_make_metavar  # type: ignore[assignment]

app = typer.Typer(add_completion=False, help="Comandos do ETL idempotente do eControle.")

@app.callback(invoke_without_command=True)
def main(ctx: typer.Context) -> None:
    """Exibe ajuda quando nenhum subcomando é informado."""

    if ctx.invoked_subcommand is None:
        typer.echo(ctx.get_help())
        raise typer.Exit()


@app.command("import")
def import_command(
    source: Path = typer.Argument(
        ...,
        exists=True,
        readable=True,
        dir_okay=False,
        resolve_path=True,
        help="Arquivo XLSM/XLSX/CSV",
    ),
    dry_run: bool = typer.Option(
        True,
        "--dry-run/--apply",
        help="Executa sem commitar alterações (use --apply para gravar)",
    ),
    file_source: Optional[str] = typer.Option(
        None,
        "--label",
        "-l",
        help="Rótulo amigável para o arquivo",
    ),
) -> None:
    run_id = str(uuid.uuid4())
    file_label = file_source or source.name

    contract = load_contract()
    raw_data = load_source(source, contract)
    normalized = transform(raw_data, contract)

    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise RuntimeError("DATABASE_URL não definido no ambiente")

    engine = sa.create_engine(database_url, future=True)
    results = run_loader(
        engine,
        normalized,
        run_id=run_id,
        file_source=file_label,
        dry_run=dry_run,
    )

    typer.echo(
        json.dumps(
            {"run_id": run_id, "dry_run": dry_run, "file_source": file_label, "results": results},
            ensure_ascii=False,
            default=str,
        )
    )


@app.command("debug-source")
def debug_source(
    source: Path = typer.Argument(
        ...,
        exists=True,
        readable=True,
        dir_okay=False,
        resolve_path=True,
        help="Arquivo XLSM/XLSX/CSV",
    ),
) -> None:
    """Mostra diagnóstico do arquivo x contrato (abas, mapeamentos e contagens extraídas)."""

    contract = load_contract()

    # 1) Abas reais do Excel
    wb = load_workbook(filename=source, read_only=True, data_only=True, keep_vba=True)
    typer.echo("\n== Abas no arquivo ==")
    for name in wb.sheetnames:
        typer.echo(f"- {name}")

    # 2) sheet_names do contrato
    typer.echo("\n== sheet_names do contrato ==")
    for logical, sheet in contract.sheet_names.items():
        typer.echo(f"{logical:20s} -> {sheet}")

    # 3) Rodar a extração e mostrar chaves e tamanhos
    raw = load_source(source, contract)

    typer.echo("\n== Seções extraídas ==")
    for key, val in raw.items():
        if isinstance(val, dict):
            # processos: dict de tabelas lógicas
            total = sum(len(v or []) for v in val.values())
            details = ", ".join(f"{k}:{len(v or [])}" for k, v in val.items())
            typer.echo(f"- {key}: {total} ({details})")
        elif isinstance(val, list):
            typer.echo(f"- {key}: {len(val)}")
        else:
            typer.echo(f"- {key}: tipo {type(val).__name__}")


if __name__ == "__main__":
    app()
