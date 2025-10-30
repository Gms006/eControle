
"""Data extraction helpers for Excel/CSV sources."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple, Optional

from openpyxl import load_workbook

from .contracts import ConfigContract
from .normalizers import normalize_text, strip_accents

ROW_NUMBER_KEY = "__row_number__"

def load(source: str | Path, contract: ConfigContract) -> Dict[str, Any]:
    path = Path(source)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return _load_excel(path, contract)
    if suffix == ".csv":
        return {"empresas": _load_csv(path)}
    raise ValueError(f"Formato não suportado: {suffix}")


def _load_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handler:
        reader = csv.DictReader(handler)
        rows: List[Dict[str, Any]] = []
        for index, row in enumerate(reader, start=2):
            row_data = {key: value for key, value in row.items()}
            row_data[ROW_NUMBER_KEY] = index
            rows.append(row_data)
    return rows


def _load_excel(path: Path, contract: ConfigContract) -> Dict[str, Any]:
    workbook = load_workbook(filename=path, read_only=False, data_only=True, keep_vba=True)
    data: Dict[str, Any] = {}
    for logical_sheet, sheet_name in contract.sheet_names.items():
        # tenta usar exatamente o nome configurado
        chosen_sheet = sheet_name if sheet_name in workbook.sheetnames else None
        # fallback: localizar por cabeçalho quando a aba não existe com o nome exato
        if chosen_sheet is None and logical_sheet == "empresas":
            chosen_sheet = _find_empresas_sheet_by_headers(workbook, contract)
        if chosen_sheet is None:
            # sem aba correspondente — pula silenciosamente (mantém compat)
            continue
        worksheet = workbook[chosen_sheet]
        tables_map = contract.table_names.get(logical_sheet, {})
        if tables_map:
            table_data = _load_tables(worksheet, tables_map)
            if table_data:
                if logical_sheet in contract.column_aliases:
                    data[logical_sheet] = _flatten_table_rows(table_data)
                else:
                    data[logical_sheet] = table_data
                continue
        data[logical_sheet] = _load_worksheet(worksheet)
    return data


def _load_tables(worksheet, tables_map: Dict[str, str]) -> Dict[str, List[Dict[str, Any]]]:
    tables: Dict[str, List[Dict[str, Any]]] = {}
    for logical_table, table_name in tables_map.items():
        # compat com diferentes versões do openpyxl
        tables_obj = getattr(worksheet, "tables", None)
        if tables_obj is None or not isinstance(tables_obj, (dict,)):
            raw_list = getattr(worksheet, "_tables", []) or []
            tables_dict = {t.name: t for t in raw_list}
        else:
            tables_dict = tables_obj
        if table_name not in tables_dict:
            continue
        table = tables_dict[table_name]
        cells = worksheet[table.ref]
        rows = _rows_from_cells(cells)
        tables[logical_table] = rows
    return tables


def _flatten_table_rows(tables: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    flattened: List[Dict[str, Any]] = []
    for rows in tables.values():
        flattened.extend(rows)
    return flattened


def _load_worksheet(worksheet) -> List[Dict[str, Any]]:
    cells = worksheet.iter_rows()
    return _rows_from_cells(cells)


def _rows_from_cells(cells: Iterable[Iterable[Any]]) -> List[Dict[str, Any]]:
    iterator = iter(cells)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []
    headers = [normalize_text(cell.value) or "" for cell in header_row]
    rows: List[Dict[str, Any]] = []
    for row in iterator:
        values = [cell.value for cell in row]
        if _is_empty_row(values):
            continue
        record: Dict[str, Any] = {}
        for header, cell in zip(headers, row):
            record[header] = cell.value
        first_cell = row[0]
        record[ROW_NUMBER_KEY] = getattr(first_cell, "row", len(rows) + 2)
        rows.append(record)
    return rows

def _find_empresas_sheet_by_headers(workbook, contract: ConfigContract) -> Optional[str]:
    """
    Quando sheet_names['empresas'] não existir, tenta localizar a aba correta
    observando a 1ª linha (cabeçalho) e checando a presença dos campos-chave:
    CNPJ, EMPRESA e MUNICIPIO conforme os aliases do config.yaml.
    """
    aliases = contract.column_aliases.get("empresas", {})
    # todos os rótulos possíveis para cada campo chave
    def _keys(key: str) -> set[str]:
        vals = aliases.get(key, [])
        return { _norm(h) for h in vals }
    need_cnpj = _keys("CNPJ")
    need_emp = _keys("EMPRESA")
    need_mun = _keys("MUNICIPIO")
    if not (need_cnpj and need_emp):
        return None
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        try:
            first = next(ws.iter_rows(max_row=1))
        except StopIteration:
            continue
        hdrs = { _norm(cell.value) for cell in first if cell.value is not None }
        # precisa bater CNPJ e EMPRESA; MUNICIPIO é desejável
        if need_cnpj & hdrs and need_emp & hdrs:
            return sheet_name
    return None

def _norm(label: Any) -> str:
    s = normalize_text(label) or ""
    # normalização forte para comparação de cabeçalho
    return strip_accents(s).casefold().replace(" ", "").replace("_", "")

def _is_empty_row(values: Iterable[Any]) -> bool:
    for value in values:
        if value not in (None, ""):
            if isinstance(value, str) and not value.strip():
                continue
            return False
    return True
